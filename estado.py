import os
import datetime
from openpyxl import load_workbook
from supabase import create_client, Client
from dotenv import load_dotenv

# Cargar variables de entorno desde el archivo .env ubicado en la carpeta actual
load_dotenv()

# Leer las variables de entorno
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    raise Exception("Las variables SUPABASE_URL y SUPABASE_KEY deben definirse en el archivo .env")

# Mostrar las variables (opcional, para confirmar que se cargaron correctamente)
print("SUPABASE_URL:", SUPABASE_URL)
print("SUPABASE_KEY:", SUPABASE_KEY)

# Crear el cliente de Supabase
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)


def transformar_fecha_detec(fecha_str):
    """
    Transforma una fecha con formato "dd/mm/yyyy HH:MM" a "YYYY-MM-DD HH:MM:SS".
    Ejemplo:
      "30/01/2025 21:42"  -> "2025-01-30 21:42:00"
    """
    try:
        dt = datetime.datetime.strptime(fecha_str, "%d/%m/%Y %H:%M")
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception as e:
        print(f"Error al transformar fecha_detec '{fecha_str}': {e}")
        return fecha_str


def leer_excel_y_preparar_registros():
    """
    Lee los archivos Excel individuales de cada empresa y prepara una lista de diccionarios.
    """
    registros = []
    # Obtener el directorio actual del script
    current_dir = os.path.dirname(os.path.abspath(__file__))
    # Construir la ruta al directorio output
    carpeta = os.path.join(current_dir, "output")

    # Verificar si la carpeta existe
    if not os.path.exists(carpeta):
        print(f"La carpeta {carpeta} no existe.")
        return registros

    archivos_excel = [
        os.path.join(carpeta, "transferencias_774691731.xlsx"),  # STS CRISTOBAL
        os.path.join(carpeta, "transferencias_777734482.xlsx"),  # DETAL
        os.path.join(carpeta, "transferencias_77936187K.xlsx")  # ST CRISTOBAL ESTADO
    ]

    for archivo in archivos_excel:
        try:
            if not os.path.exists(archivo):
                print(f"El archivo {archivo} no existe, continuando con el siguiente...")
                continue

            print(f"Leyendo archivo: {archivo}")
            wb = load_workbook(archivo)
            ws = wb.active

            # Iterar desde la segunda fila (omitiendo la cabecera)
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Transformar la fecha de detección
                fecha_detec_original = row[1]  # Ejemplo: "30/01/2025 21:42"
                fecha_detec_formateada = transformar_fecha_detec(fecha_detec_original) if fecha_detec_original else None

                registro = {
                    "hash": row[8],
                    "empresa": row[7],
                    "monto": row[5],
                    "rut": row[4],
                    "rs": row[3],
                    "fecha": row[2],
                    "fecha_detec": fecha_detec_formateada,
                    "facturación": row[6],
                    "enviada": "0"  # Valor predeterminado para la columna 'enviada'
                }
                registros.append(registro)

            print(f"Se leyeron {ws.max_row - 1} registros del archivo {archivo}")

        except Exception as e:
            print(f"Error al leer el archivo {archivo}: {e}")
            continue

    return registros


def subir_registros_a_supabase(registros):
    """
    Inserta los registros en la tabla 'transferencias' de Supabase.
    Si se detecta un registro duplicado en base al campo 'hash', se ignora.
    """
    registros_insertados = 0
    registros_duplicados = 0
    registros_con_error = 0

    for registro in registros:
        try:
            response = supabase.table("transferencias").insert(registro).execute()
            registros_insertados += 1
            print(f"Registro insertado: {registro['hash']}")
        except Exception as e:
            # Si el error indica duplicidad, se ignora ese registro
            if "duplicate key value violates unique constraint" in str(e):
                registros_duplicados += 1
                print(f"Registro duplicado ignorado: {registro['hash']}")
            else:
                registros_con_error += 1
                print(f"Error al insertar registro {registro['hash']}: {e}")

    print(f"\nResumen de la subida a Supabase:")
    print(f"Registros insertados exitosamente: {registros_insertados}")
    print(f"Registros duplicados ignorados: {registros_duplicados}")
    print(f"Registros con error: {registros_con_error}")


def main():
    print("Iniciando proceso de subida a Supabase...")
    registros = leer_excel_y_preparar_registros()
    print(f"Se encontraron {len(registros)} registros para subir.")

    if registros:
        subir_registros_a_supabase(registros)
    else:
        print("No se encontraron registros para subir.")


if __name__ == "__main__":
    main()
